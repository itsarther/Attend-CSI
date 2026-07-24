import io
import csv
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from app.models.attendance import Attendance
from app.models.event import Event


def export_to_csv(event: Event, records: List[Attendance]) -> io.BytesIO:
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header Info
    writer.writerow([f"Event Attendance Record: {event.name}"])
    writer.writerow([f"Organized By: {event.organized_by}", f"Date: {event.date}", f"Venue: {event.venue}"])
    writer.writerow([])
    
    # Table Header
    writer.writerow([
        "Sr No", "Student Name", "GR Number", "Roll Number", "Department",
        "Year", "Semester", "Class", "Division", "Mobile",
        "Submission Time", "Method", "Verification Status", "IP Address"
    ])
    
    for idx, rec in enumerate(records, start=1):
        writer.writerow([
            idx,
            rec.student_name,
            rec.gr_number,
            rec.roll_number,
            rec.department,
            rec.year,
            rec.semester,
            rec.class_name,
            rec.division,
            rec.mobile or "N/A",
            rec.submission_time.strftime("%Y-%m-%d %H:%M:%S") if rec.submission_time else "",
            rec.submission_method,
            rec.verification_status,
            rec.ip_address or "N/A"
        ])
    
    mem_buf = io.BytesIO()
    mem_buf.write(output.getvalue().encode('utf-8'))
    mem_buf.seek(0)
    return mem_buf


def export_to_excel(event: Event, records: List[Attendance]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Roster"

    # Styling definitions
    title_font = Font(name="Calibri", size=16, bold=True, color="1E293B")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="475569")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # CSI Indigo Blue
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    border_side = Side(style='thin', color='E2E8F0')
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Title Block
    ws.merge_cells("A1:N1")
    ws["A1"] = f"Attend | CSI - Official Attendance Roster: {event.name}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:N2")
    ws["A2"] = f"Organized By: {event.organized_by} | Venue: {event.venue} | Date: {event.date} | Total Registered: {len(records)}"
    ws["A2"].font = subtitle_font

    ws.append([]) # Empty row 3

    # Table Header Row 4
    headers = [
        "Sr No", "Student Name", "GR Number", "Roll Number", "Department",
        "Year", "Semester", "Class", "Division", "Mobile",
        "Submission Time", "Method", "Status", "IP Address"
    ]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    # Data Rows
    for idx, rec in enumerate(records, start=1):
        row_num = idx + 4
        row_data = [
            idx,
            rec.student_name,
            rec.gr_number,
            rec.roll_number,
            rec.department,
            rec.year,
            rec.semester,
            rec.class_name,
            rec.division,
            rec.mobile or "N/A",
            rec.submission_time.strftime("%Y-%m-%d %H:%M:%S") if rec.submission_time else "",
            rec.submission_method,
            rec.verification_status,
            rec.ip_address or "N/A"
        ]
        ws.append(row_data)

        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = cell_border
            if idx % 2 == 0:
                cell.fill = alt_row_fill
            if col_num in [1, 3, 4, 6, 7, 8, 9, 12, 13]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column Auto-Width adjustment
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    mem_buf = io.BytesIO()
    wb.save(mem_buf)
    mem_buf.seek(0)
    return mem_buf


def export_to_pdf(event: Event, records: List[Attendance]) -> io.BytesIO:
    mem_buf = io.BytesIO()
    doc = SimpleDocTemplate(
        mem_buf,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=6
    )
    meta_style = ParagraphStyle(
        'DocMeta',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#475569'),
        spaceAfter=15
    )
    table_text_style = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontSize=8,
        leading=10
    )
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.white,
        fontName='Helvetica-Bold'
    )

    elements = []
    elements.append(Paragraph(f"<b>Attend | CSI</b> - Event Attendance Report: {event.name}", title_style))
    elements.append(Paragraph(f"Organized By: {event.organized_by} &nbsp;|&nbsp; Venue: {event.venue} &nbsp;|&nbsp; Date: {event.date} &nbsp;|&nbsp; Total Attendance: {len(records)}", meta_style))

    # Table Header
    headers = ["Sr", "Student Name", "GR No", "Roll", "Dept", "Yr", "Sem", "Div", "Time", "Method", "Status"]
    table_data = [[Paragraph(h, table_header_style) for h in headers]]

    for idx, rec in enumerate(records, start=1):
        t_str = rec.submission_time.strftime("%H:%M:%S") if rec.submission_time else ""
        row = [
            Paragraph(str(idx), table_text_style),
            Paragraph(rec.student_name, table_text_style),
            Paragraph(rec.gr_number, table_text_style),
            Paragraph(rec.roll_number, table_text_style),
            Paragraph(rec.department, table_text_style),
            Paragraph(rec.year, table_text_style),
            Paragraph(str(rec.semester), table_text_style),
            Paragraph(rec.division, table_text_style),
            Paragraph(t_str, table_text_style),
            Paragraph(rec.submission_method, table_text_style),
            Paragraph(rec.verification_status, table_text_style),
        ]
        table_data.append(row)

    col_widths = [25, 120, 60, 40, 110, 30, 30, 35, 60, 75, 65]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))

    elements.append(t)
    doc.build(elements)
    mem_buf.seek(0)
    return mem_buf
